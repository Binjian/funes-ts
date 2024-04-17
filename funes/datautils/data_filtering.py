#!/usr/bin/env python
# coding: utf-8

# This script is to filter the data downloaded by getData.ipynb
# required data in json file
# will store filtered data into .xlsx

# In[13]:

# system packages
import json
import time

# third party packages
import openpyxl


def main():
    # define data-fetch file
    wb = openpyxl.Workbook()
    carno = "HMZABAAH9MF014494"
    ws = wb.create_sheet(carno)
    # data json file
    filename = "../../data/all-info/"+carno+".json"
    # load json
    with open(filename) as file_obj:
        res = json.load(file_obj)

    # the code below only filter the data into time,current and voltage1-198

    # start from second row
    num = 2

    for keys in res:
        # save cell voltage
        Vol = []
        for i in range(1, 10):
            try:
                Vol.append(keys["_source"]["canData"]["data"]["CellVoltage_0" + str(i)])
            except:
                print(str(i) + "th dose not exist")
                continue
        for i in range(10, 199):
            try:
                Vol.append(keys["_source"]["canData"]["data"]["CellVoltage_" + str(i)])
            except:
                print(str(i) + "th dose not exist")
        if sum(Vol) != 0 and Vol.count(8191) == 0:  # 8191 is error code.
            vnum = 1
            for v in Vol:
                ws.cell(row=num, column=vnum + 3).value = v
                # iterate all cellvoltage_i
                vnum += 1

            try:
                # save time, current and soc
                ws.cell(row=num, column=1).value = time.strftime(
                    "%Y-%m-%d %H:%M:%S",
                    time.localtime(int(keys["_source"]["travelTime"]) / 1000),
                )
                ws.cell(row=num, column=2).value = keys["_source"]["canData"]["data"][
                    "VCU_BMSPackCurrent"
                ]
                ws.cell(row=num, column=3).value = keys["_source"]["canData"]["data"][
                    "BMSSOC"
                ]
            except:
                continue
            num += 1  # increment of row
        else:
            print("8191 error!")

    # title "CellVoltage1-198"
    for i in range(1, 199):
        ws.cell(row=1, column=i + 3).value = "CellVoltage_" + str(i)
    # time, current and soc to file
    ws.cell(row=1, column=1).value = "time"
    ws.cell(row=1, column=2).value = "current"
    ws.cell(row=1, column=3).value = "BMSSOC"

    # save in xlsx
    wb.save("filt-data/NewCar4.xlsx")
    print("success")


if __name__ == "__main__":
    main()
