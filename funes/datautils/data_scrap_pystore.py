#!/usr/bin/env python
# coding: utf-8

# This script is to download data from elastic platform by some features(car number,time range)
# data will be store in json file
import os
import traceback
from typing import MutableSet
from elasticsearch import Elasticsearch
import json
from collections import Counter
import openpyxl
import json
import time
from collections import Counter
import openpyxl
from scipy.interpolate import interp1d
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
import logging
import pystore
import warnings
warnings.filterwarnings('ignore')

### data fetch

car_list = ["HMZABAAH6MF014484",
"HMZABAAH8MF014485",
"HMZABAAH4MF014483",
"HMZABAAH2MF014482",
"HMZABAAH1MF014487",
"HMZABAAH5MF014489",
"HMZABAAH5MF014492",
"HMZABAAH3MF014491",
"HMZABAAH1MF014490",
"HMZABAAH9MF014494",
"HMZABAAH6MF014498",
"HMZABAAH8MF014499",
"HMZABAAH2MF014501",
"HMZABAAH3MF018296",
"HMZABAAH1MF018295",
"HMZABAAH4MF018274",
"HMZABAAHXMF018294",
"HMZABAAH4MF018288",
"HMZABAAH2MF018290",
"HMZABAAHXMF018277",
"HMZABAAH6MF018289",
"HMZABAAH9MF018285",
"HMZABAAH8MF018276",
"HMZABAAH4MF014497",
"HMZABAAH0MF014495",
"HMZABAAH4MF014502",
"HMZABAAH6MF014503",
"HMZABAAH2MF018306",
"HMZABAAH6MF018308",
"HMZABAAH4MF018307",
"HMZABAAH0MF018305",
"HMZABAAH8MF018312",
"HMZABAAH7MF018298",
"HMZABAAH1MF014506",
"HMZABAAH4MF018310",
"HMZABAAH6MF018275",
"HMZABAAH2MF018273",
"HMZABAAH7MF018317",
"HMZABAAH9MF018335",
"HMZABAAH0MF014500",
"HMZABAAH3MF014488",
"HMZABAAH2MF018323",
"HMZABAAH1MF018314",
"HMZABAAH4MF018341",
"HMZABAAH3MF018279",
"HMZABAAH7MF018320",
"HMZABAAHXMF018330",
"HMZABAAH3MF018332",
"HMZABAAH9MF018352",
"HMZABAAH2MF018340",
"HMZABAAH6MF018339",
"HMZABAAH4MF018291",
"HMZABAAH2MF018368",
"HMZABAAH3MF018282",
"HMZABAAH3MF018301",
"HMZABAAH0MF018272",
"HMZABAAH8MF018293",
"HMZABAAH6MF018292",
"HMZABAAHXMF018327",
"HMZABAAH5MF018350",
"HMZABAAH3MF014507",
"HMZABAAH7MF018348",
"HMZABAAH9MF018349",
"HMZABAAH9MF018318",
"HMZABAAH0MF018286",
"HMZABAAH6MF018342",
"HMZABAAH6MF018325",]

# define unix timestamp
day_timestamp = 86400000
week_timestamp = day_timestamp * 7

start_scrap = datetime.now().strftime("%Y-%m-%d")

# car vin number to search
for car in car_list:
    print("current vim: ", car)
    carNo = car + datetime.now().strftime("%Y-%m-%d-%H-%M-%S")

    # time range
    now = datetime.now().strftime("%Y.%m.%d")
    if (int(now.split('.')[-1]) - 7)!=0:
        if (int(now.split('.')[-1]) - 7) < 10:
            start_time = '.'.join(now.split('.')[:-1]) + '.0' + str(int(now.split('.')[-1]) - 7) + ' 00:00:00'
        else:
            start_time = '.'.join(now.split('.')[:-1]) + '.' + str(int(now.split('.')[-1]) - 7) + ' 00:00:00'
    else:
        start_time = '.'.join(now.split('.')[:-1]) + '.0' + str(int(now.split('.')[-1]) - 6) + ' 00:00:00'
    end_time = now + ' 00:00:00'
    print('###############################################')
    print(f'data scrapping from {start_time} to {end_time}')
    print('###############################################')
    # start_time = "2022.6.24 00:00:00"
    # end_time = "2022.10.24 00:00:00"
    # convert time range to timestamp
    start_timestamp = int(
        datetime.strptime(start_time, "%Y.%m.%d %H:%M:%S").timestamp() * 1000
    )
    end_timestamp = int(datetime.strptime(end_time, "%Y.%m.%d %H:%M:%S").timestamp() * 1000)
    # define time gap
    time_gap = end_timestamp - start_timestamp
    cycle = int(time_gap / week_timestamp)
    init_res = []

    for i in range(cycle):
        start_time = start_timestamp + i * week_timestamp
        end_time = start_time + week_timestamp
        print(datetime.fromtimestamp(start_time / 1000))
        print(datetime.fromtimestamp(end_time / 1000))
        # define soc range
        start_soc = 80
        end_soc = 100
        # define elastic address & port and authentication
        es = Elasticsearch(
            hosts=[
                "http://10.0.80.13:9200",
                "http://10.0.80.14:9200",
                "http://10.0.80.15:9200",
            ],
            basic_auth=("elastic", "123456"),
        )
        # start fetch
        print("start fetch data")
        # define search range and filter
        body = {
            "query": {
                "bool": {
                    # must match car vin number and charge state
                    "must": [
                        {"match": {"vin": car}},
                        {"match": {"canData.data.VCU_BMSChargeState": 3}},
                    ],
                    # filt the search time range
                    "filter": [
                        {
                            "range": {
                                "travelTime": {
                                    "gte": start_time,
                                    "lte": end_time,
                                }
                            }
                        },
                        {
                            "range": {
                                "canData.data.BMSSOC": {
                                    "gte": start_soc,
                                    "lte": end_soc,
                                }
                            }
                        },
                    ],
                }
            },
            # sort data via travel time and by order of asc
            "sort": [{"travelTime": {"order": "asc"}}],
        }
        # index is the project name, size is defined size of fetched data
        try:
            res = es.options().search(
                index="tsp-vehicle-log3*", body=body, size=20000, request_timeout=60
            )
            # print('########',res)
            if i == 0:
                init_res = res["hits"]["hits"]
            else:
                init_res.extend(res["hits"]["hits"])
            # print('######################',len(init_res),res)
            print("week search done")
        except Exception as e:
            print(e)
            print("elastic_transport.ConnectionTimeout: Connection timed out")
            print('week search break. next week.')
            continue

    # print(res)
    # print(init_res)
    # with open(filename, "w") as file_obj:
    #     json.dump(init_res, file_obj)
    print("end fetch data")

    # data-filtering
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(carNo)
    # data json file
    # filename = "../../data/all-info/" + carNo + ".json"
    # load json
    # with open(filename) as file_obj:
    #     res = json.load(file_obj)

    res = init_res

    soc = [
        0,
        2,
        5,
        10,
        15,
        20,
        25,
        30,
        35,
        40,
        45,
        50,
        55,
        60,
        65,
        70,
        75,
        80,
        85,
        90,
        95,
        98,
        100,
    ]
    ocv = [
        2737.5,
        3085.5,
        3180,
        3209.3,
        3223.1,
        3250.4,
        3266.7,
        3282.1,
        3289.2,
        3289.9,
        3290.6,
        3291.3,
        3293.2,
        3304.1,
        3329.6,
        3330.3,
        3331,
        3332,
        3332.9,
        3334.1,
        3334.5,
        3365,
        3443.2,
    ]

    get_ocv = interp1d(soc, ocv, kind="linear")


    def ISOformat(times):
        t = datetime(times[0], times[1], times[2], times[3], times[4], times[5])
        return t.isoformat()


    # start from second row
    num = 2
    try:
        for keys in res:
            # save cell voltage
            Vol = []
            for i in range(1, 10):
                Vol.append(keys["_source"]["canData"]["data"]["CellVoltage_0" + str(i)])
                continue
            for i in range(10, 199):
                Vol.append(keys["_source"]["canData"]["data"]["CellVoltage_" + str(i)])

            if sum(Vol) != 0:
                vnum = 1
                for v in Vol:
                    ws.cell(row=num, column=vnum + 4).value = v
                    # iterate all cellvoltage_i
                    vnum += 1
                # print(ws.values)

                # save time, current and soc
                source_time = ISOformat(time.localtime(int(keys["_source"]["travelTime"]) / 1000)[:6])
                # ws.cell(row=num, column=1).value = time.strftime(
                #     "%Y-%m-%d %H:%M:%S",
                #     time.localtime(int(keys["_source"]["travelTime"]) / 1000)
                # )
                ws.cell(row=num, column=1).value = source_time
                # print('###########',keys["_source"]["travelTime"])
                ws.cell(row=num, column=2).value = keys["_source"]["canData"]["data"][
                    "VCU_BMSPackCurrent"
                ]
                ws.cell(row=num, column=3).value = keys["_source"]["canData"]["data"]["BMSSOC"]
                ocv = get_ocv(keys["_source"]["canData"]["data"]["BMSSOC"])
                ws.cell(row=num, column=4).value = float(ocv)

                num += 1  # increment of row
            else:
                print("error!")

        # title "CellVoltage1-198"
        for i in range(1, 199):
            ws.cell(row=1, column=i + 4).value = "CellVoltage_" + str(i)
        # time, current and soc to file
        ws.cell(row=1, column=1).value = "time"
        ws.cell(row=1, column=2).value = "current"
        ws.cell(row=1, column=3).value = "BMSSOC"
        ws.cell(row=1, column=4).value = "ocv"

        print("data filtering success")


        # data processing

        def cutSegment(size, data, segmt):
            segdata = pd.DataFrame()
            for i in range(len(segmt) - 1):
                length = segmt[i + 1] - segmt[i]
                part = pd.DataFrame()
                if length > size:
                    n = length - size
                    part = data.iloc[segmt[i]: (segmt[i + 1]), :]
                    part = part.reset_index(drop=True)
                    part.drop(part.tail(n).index, inplace=True)
                    segdata = segdata.append(part, ignore_index=True)
            return segdata


        # read xlsx file with # time, current, soc and vol1-198

        # store data to data frame
        df = pd.DataFrame()

        for i in ws.values:
            for j in i:
                df[j] = []
            break

        for i, v in enumerate(list(ws.columns)):
            df[list(df)[i]] = [j.value for j in v][1:]

        ######################################################################
        ## data processing
        # create new data frame to store processed data
        centerdf = pd.DataFrame()
        # variable number
        n = 4
        # store time, current and soc to data frame
        centerdf["time"] = df.iloc[:, 0]
        centerdf["current"] = df.iloc[:, 1]
        centerdf["soc"] = df.iloc[:, 2]
        centerdf["ocv"] = df.iloc[:, 3]

        # store difference between mean of cell voltage for each cell into data frame
        for i in range((int(df.shape[1] - n))):
            # print(df.shape[1] - n,i + 1)
            centerdf["vol" + str(i + 1)] = df.iloc[:, i + n]

        # separate data by time
        seg = [0]
        for i in range(0, (int(df.shape[0]) - 1)):
            if df.iloc[i + 1, 3] == 3:
                if datetime.strptime(
                        df.iloc[i + 1, 0], "%Y-%m-%d %H:%M:%S"
                ) - datetime.strptime(df.iloc[i, 0], "%Y-%m-%d %H:%M:%S") >= timedelta(
                    seconds=300
                ):
                    seg.append(i + 1)
        seg.append(len(df))
        # print(seg)
        # insert first value as default to fill the segment
        segdf = cutSegment(370, centerdf, seg)
        # print(segdf)
        # save processed data into csv format for further training
        vol = []
        cur = []
        soc = []
        ocv = []
        timestamp = []

        store = pystore.store('train_data')
        # store.delete_collection('2022-08-09')
        # store.delete_collection('test')
        # collection = store.collection(datetime.now().strftime("%Y-%m-%d"))
        collection = store.collection(start_scrap)

        sheet0 = pd.DataFrame()
        sheet0[''] = ['unit', 'description', 'range', 'type']
        sheet0['timestamp'] = ['s', 'unix timestamp', '[2022.01.01,inf)', 'float']
        sheet0['current'] = ['A', 'battery current', 'None', 'float']
        sheet0['vol'] = ['dumV', 'cell voltage', 'None', 'float']
        sheet0['soc'] = ['percentage', 'None', 'None', 'float']
        sheet0['ocv'] = ['dumV', 'None', 'None', 'float']
        collection.write('metadata', sheet0, metadata={'source': 'metadata'}, overwrite=True)

        for i in range(198):
            vol = segdf.iloc[:, i + 4].values.tolist()
            timestamp = segdf.iloc[:, 0].values.tolist()
            # print(timestamp)
            # timestamp = [np.float64(time.mktime(time.strptime(j, '%Y-%m-%d %H:%M:%S'))) for j in timestamp]
            cur = segdf.iloc[:, 1].values.tolist()
            soc = segdf.iloc[:, 2].values.tolist()
            ocv = segdf.iloc[:, 3].values.tolist()

            tsdf = pd.DataFrame()
            tsdf["timestamp"] = pd.to_datetime(timestamp)
            tsdf["current"] = np.float64(cur)
            tsdf["vol"] = np.float64(vol)
            tsdf["soc"] = np.float64(soc)
            tsdf["ocv"] = np.float64(ocv)
            collection.write(carNo[:17] + '_cell_' + str(i + 1), tsdf, metadata={'source': carNo[:17]})
        print(len(collection.list_items()))
        print(len(list(set([j[:17] for j in [i.split('_')[0] for i in list(collection.list_items())
                                             if i != 'metadata']]))))
    except Exception as e:
        print(e)
        print(f'drop {carNo}')
        continue


